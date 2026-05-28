import pandas as pd
from openpyxl import load_workbook  # Import load_workbook


def validate_teach_login(username, password):
    # Load the Excel workbook
   
    workbook = load_workbook('data/teach_log.xlsx')
    sheet = workbook.active

 # Iterate through rows in the Excel sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #print("Row:", row)
        stored_username = str(row[0]).strip()
        stored_password = str(row[1]).strip()
        teacher_name = str(row[2]).strip()  # Assuming teacher name is in the third column
        user_type = str(row[3]).strip()  # Assuming user type is in the fourth column


        if username == stored_username and password == stored_password:
            return {'username': stored_username, 'teacher_name': teacher_name, 'user_type': user_type}

    return None