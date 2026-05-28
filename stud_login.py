from openpyxl import load_workbook  # Import load_workbook

def validate_stud_login(enrollment, password):
    # Load the Excel workbook
    # Replace 'user_data.xlsx' with the actual path to your Excel file
    workbook = load_workbook('data/users.xlsx')
    sheet = workbook.active

    
    # Iterate through rows in the Excel sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stored_enrollment = str(row[0]).strip()
        stored_password = str(row[1]).strip()
        student_name = str(row[2]).strip()  # Assuming student name is in the third column

        if enrollment == stored_enrollment and password == stored_password:
            return {'enrollment': stored_enrollment, 'student_name': student_name}

    return None

